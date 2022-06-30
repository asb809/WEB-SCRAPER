from bs4 import BeautifulSoup
import pandas as pd
import requests
import csv
from openpyxl import load_workbook
import matplotlib.pyplot as plt
from tkinter import *
from collections import Counter

def DataScrap():
    html = requests.get('https://www.bbc.com/urdu')
    content = html.content
    soup = BeautifulSoup(content, 'html.parser')
    categories = soup.find_all('a', class_='bbc-puhg0e e1ibkbh73')
    category_name = []
    links = []

    # print(categories['href'])
    for link in soup.findAll('a', {'class': 'bbc-puhg0e e1ibkbh73'}):
        try:
            links.append(link['href'])
            category_name.append(link.text)
        except KeyError:
            pass
    links = links[:-1]
    # print(links)
    # print(category_name)


    dic = {}
    stories_headlines = []
    stories_links = []
    for i in range(len(links)):

        category = f'Category: ' + categories[i].text

        headline = []
        story_link = []

        for pagenum in range(1, 6):

            if pagenum == 1:
                curr_link = 'https://www.bbc.com' + links[i]
            #    print(curr_link)
            else:
                curr_link = 'https://www.bbc.com' + links[i] + "?page=" + str(pagenum)
                # print(curr_link)
            print(curr_link)
            html = requests.get(curr_link)
            content = html.content
            soup = BeautifulSoup(content, 'html.parser')
            link_categories = soup.find_all('a', class_='bbc-uk8dsi emimjbx0')

            for x in link_categories:
                #  print(f'Category: '+ category_name[i])
                #   print(f'Link desc: '+ x.text)
                headline.append(x.text)
                #    print(f'Link: '+x['href'])
                story_link.append(x['href'])

        stories_headlines.append(headline[:100])
        stories_links.append(story_link[:100])

    stories = []
    i = 1
    for story in stories_links:
        story_text = []
        for j in range(100):
            # print(j, ": ", story[j])
            html = requests.get(story[j])
            content = html.content
            soup = BeautifulSoup(content, 'html.parser')
            link_categories = soup.find_all('p', class_='bbc-yabuuk e1cc2ql70')
            lines = []
            for b in link_categories:
                lines.append(b.text)
            storyyy = '\n'.join(lines)
            story_text.append(storyyy)
        stories.append(story_text)

    cat1 = []
    cat2 = []
    cat3 = []
    cat4 = []
    cat5 = []
    cat6 = []
    for i in range(100):
        cat1.append(category_name[0])
        cat2.append(category_name[1])
        cat3.append(category_name[2])
        cat4.append(category_name[3])
        cat5.append(category_name[4])
        cat6.append(category_name[5])

    df = pd.DataFrame(zip(stories[0], stories_headlines[0], cat1), columns=['Story', 'Headline', 'Category'])
    df1 = pd.DataFrame(zip(stories[1], stories_headlines[1], cat2), columns=['Story', 'Headline', 'Category'])
    df2 = pd.DataFrame(zip(stories[2], stories_headlines[2], cat3), columns=['Story', 'Headline', 'Category'])
    df3 = pd.DataFrame(zip(stories[3], stories_headlines[3], cat4), columns=['Story', 'Headline', 'Category'])
    df4 = pd.DataFrame(zip(stories[4], stories_headlines[4], cat5), columns=['Story', 'Headline', 'Category'])
    df5 = pd.DataFrame(zip(stories[5], stories_headlines[5], cat6), columns=['Story', 'Headline', 'Category'])

    df = df.append(df1, ignore_index=True)
    df = df.append(df2, ignore_index=True)
    df = df.append(df3, ignore_index=True)
    df = df.append(df4, ignore_index=True)
    df = df.append(df5, ignore_index=True)

    # crearing excel file
    df.to_excel(r'BBC.xlsx', index=False)



def uniqueWords():
    data = 'BBC.xlsx'
    work_book = load_workbook(data)
    work_sheet = work_book['Sheet1']
    all_columns = list(work_sheet.columns)
    all_stories = ''
    for i in all_columns[0]:
        if(i.value != 'Story'):
            all_stories = all_stories + str(i.value)
    new_allStories = ''
    for i in all_stories:
        if i.isalnum() or i.isspace():
            new_allStories = new_allStories + i
    all_words = new_allStories.split()
    List = []
    for i in all_words:
        if i not in List:
            List.append(i)

    # i = 0
    # while i < len(List):
    #     # print(List[i])
    #     i+=1

    a = Tk()
    a.geometry('400x400')
    text = Text(a, height=20, width=40)
    text.pack()
    text.insert(END, 'Total Unique words are: '+str(len(List)))
    a.mainloop()


def maxLengthStory():
    data = 'BBC.xlsx'
    work_book = load_workbook(data)
    work_sheet = work_book['Sheet1']
    all_columns = list(work_sheet.columns)
    max = 0
    maxStory = ''
    for i in all_columns[0]:
        if(i.value != 'Story'):
            if max < len(str(i.value)):
                max = len(str(i.value))
                maxStory = str(i.value)
    a = Tk()
    a.geometry('400x400')
    text = Text(a, height=20, width=40)
    text.pack()
    text.insert(END, maxStory+'\n'+'Length of story: '+str(len(maxStory)))
    a.mainloop()

def minLengthStory():
    data = 'BBC.xlsx'
    work_book = load_workbook(data)
    work_sheet = work_book['Sheet1']
    all_columns = list(work_sheet.columns)
    min = 2**31-1
    minStory = ''
    for i in all_columns[0]:
        if i.value != 'Story' and i.value != None:
            if min > len(str(i.value)):
                min = len(str(i.value))
                minStory = str(i.value)

    a = Tk()
    a.geometry('400x400')
    text = Text(a,height=20,width=40)
    text.pack()
    text.insert(END,minStory+'\n'+'Length of story: '+str(len(minStory)))
    a.mainloop()

def topfrequency():
    data = 'BBC.xlsx'
    work_book = load_workbook(data)
    work_sheet = work_book['Sheet1']
    all_columns = list(work_sheet.columns)
    all_stories = ''
    for i in all_columns[0]:
        if (i.value != 'Story'):
            all_stories = all_stories + str(i.value)
    str1 = ''
    for i in all_stories:
        if i.isalnum() or i.isspace():
            str1 = str1 + i
    arr_words = str1.split()
    counter = Counter(arr_words)
    sen = counter.most_common(10)
    a = Tk()
    a.geometry('400x400')
    text = Text(a, height=20, width=40)
    text.pack()
    text.insert(END, '\n' + 'Top Frequency: \n' + str(sen))
    a.mainloop()

    # print(sen)

def bar_graph():
    blogs =['Pakistan','Aas Paas','World','Sports','Entertainment','Science']
    Stories = [100,100,100,100,100,100]

# Creating a simple bar chart
    plt.bar(blogs, Stories)

    plt.title('Number Of Stories From Different Categories')
    plt.xlabel('Categories', fontsize=15)
    plt.ylabel('posts', fontsize=15)
    plt.show()



# bar_graph()
# maxLengthStory()
# minLengthStory()
# uniqueWords()
# topfrequency()


mw = Tk()
mw.geometry('800x768+0+0')
mw.resizable(False, False)
frame = Frame(mw)
frame.pack()
#Main GUI Interface
label = Label(text='ENTER WEBSITE LINK',font=('Arial 16 bold'),bg='#34495E',fg='white', borderwidth=7)
label.place(x=10,y=25)

url = Entry(mw, font=('Arial 16 bold'), width=30, borderwidth=7)
url.place(x=270, y=25)

button = Button(text='SCRAP',bg='#27AE60',fg='white',command=DataScrap,font=('Arial 16 bold'), borderwidth=7)
button.place(x=660,y=25, relheight=0.05)

label1 = Label(text='SHORTEST STORY',font=('Arial 16 bold'),bg='#34495E',fg='white',  borderwidth=7)
label1.place(x=320,y=100)
button1 = Button(text='Click Here',bg='#27AE60',fg='white',command=minLengthStory,font=('Arial 16 bold'), borderwidth=5)
button1.place(x=360,y=150)

label2 = Label(text='LONGEST STORY',font=('Arial 16 bold'),bg='#34495E',fg='white', borderwidth=7)
label2.place(x=330,y=220)
button2 = Button(text='Click Here',bg='#27AE60',fg='white',command=maxLengthStory,font=('Arial 16 bold'), borderwidth=5)
button2.place(x=360,y=270)

label3 = Label(text='WORDS BY FREQUENCY',font=('Arial 16 bold'),bg='#34495E',fg='white',  borderwidth=7)
label3.place(x=295,y=340)
button3 = Button(text='Click Here',bg='#27AE60',fg='white',command=topfrequency,font=('Arial 16 bold'),  borderwidth=5)
button3.place(x=360,y=390)

# label4 = Label(text='STORIES IN EACH CATEGORY',font=('Arial 16 bold'),bg='#34495E',fg='white', borderwidth=7)
# label4.place(x=270,y=460)
# button4 = Button(text='Click Here',bg='#27AE60',fg='white',font=('Arial 16 bold'),borderwidth=5)
# button4.place(x=360,y=510)

label5 = Label(text='UNIQUE WORDS',font=('Arial 16 bold'),bg='#34495E',fg='white', borderwidth=7)
label5.place(x=175,y=580)
button5 = Button(text='Click Here',bg='#27AE60',fg='white',command=uniqueWords,font=('Arial 16 bold'), borderwidth=5)
button5.place(x=200,y=630)

label6 = Label(text='PLOT GRAPH',font=('Arial 16 bold'),bg='#34495E',fg='white', borderwidth=7)
label6.place(x=500,y=580)
button6 = Button(text='Click Here',bg='#27AE60',fg='white',command=bar_graph,font=('Arial 16 bold'), borderwidth=5)
button6.place(x=513,y=630)

mw.mainloop()
